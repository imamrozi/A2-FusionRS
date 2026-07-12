"""
generate_excel_report.py

Susun laporan Excel bertahap (dasar -> detail) dari hasil eksperimen A2-IRM
untuk presentasi ke pembimbing akademik (promotor). Sumber data:
checkpoints/results/ (manuscript_table_*.csv, YAML per-seed, predictions
per-seed CSV) -- TIDAK menjalankan eksperimen baru, murni menyusun ulang
data yang sudah tervalidasi (lihat manuscript/A2-IRM_manuscript_draft.md).

8 sheet, 2 tingkatan:
  - Tier "Wajib" (Sheet 1-3): ringkasan eksekutif, hasil utama RMSE/MAE,
    perbandingan hybrid vs CF klasik -- cukup dibaca utk memahami kesimpulan.
  - Tier "Lampiran" (Sheet 4-7): signifikansi statistik, stabilitas/varian,
    data mentah per-seed, karakteristik dataset & hyperparameter -- dibuka
    kalau ingin menelusuri detail/verifikasi.

Usage:
    python generate_excel_report.py \
        --results-dir checkpoints/results \
        --output manuscript/A2-IRM_hasil_eksperimen.xlsx
"""

from __future__ import annotations

import argparse
import logging
from datetime import date
from pathlib import Path

import numpy as np
import pandas as pd
import yaml
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.evaluation.metrics import significance_test

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

DOMAIN_ORDER = ["restaurant", "amazon_electronics", "tripadvisor_hotel"]
DOMAIN_LABEL = {
    "restaurant": "Restoran (Yelp)",
    "amazon_electronics": "E-commerce (Amazon Electronics)",
    "tripadvisor_hotel": "Hotel (TripAdvisor)",
}
MODEL_ORDER = [
    "baseline_reimpl",
    "absa_ablation",
    "absa_ablation_confidence_mean",
    "absa_ablation_concat",
    "absa_ablation_concat_confidence",
    "classical_cf_item_knn",
    "classical_cf_svd",
]
MODEL_LABEL = {
    "baseline_reimpl": "Baseline (Reimplementasi Darraz et al.)",
    "absa_ablation": "ABSA - Mean",
    "absa_ablation_confidence_mean": "ABSA - Confidence-Weighted Mean",
    "absa_ablation_concat": "ABSA - Concat",
    "absa_ablation_concat_confidence": "ABSA - Concat + Confidence",
    "classical_cf_item_knn": "CF Klasik - Item-KNN",
    "classical_cf_svd": "CF Klasik - SVD",
}
CORE_MODELS = MODEL_ORDER[:5]  # baseline + 4 varian ABSA (tanpa classical CF)
ABSA_MODELS = MODEL_ORDER[1:5]
CF_MODELS = ["classical_cf_item_knn", "classical_cf_svd"]
COMPARISON_MODELS = ABSA_MODELS + CF_MODELS  # dibandingkan vs baseline_reimpl
SEEDS = [42, 123, 456, 789, 1011]

DATASET_STATS = {
    "restaurant": dict(reviews=118695, users=7152, items=3757, sparsity=0.9956, rating=3.76, test_size=13233),
    "amazon_electronics": dict(reviews=122068, users=14750, items=9226, sparsity=0.9991, rating=4.37, test_size=16580),
    "tripadvisor_hotel": dict(reviews=79562, users=11236, items=2056, sparsity=0.9966, rating=3.94, test_size=11795),
}

FILL_HEADER = "1F4E78"
FILL_TIER_WAJIB = "1F6F45"
FILL_TIER_LAMPIRAN = "6E6E6E"
FILL_GREEN_SIG = "C6EFCE"


# ===========================================================================
# Data loaders
# ===========================================================================

def load_manuscript_tables(results_dir: Path) -> pd.DataFrame:
    dfs = [pd.read_csv(results_dir / f"manuscript_table_{d}.csv") for d in DOMAIN_ORDER]
    df = pd.concat(dfs, ignore_index=True)
    df["domain"] = pd.Categorical(df["domain"], categories=DOMAIN_ORDER, ordered=True)
    df["model_prefix"] = pd.Categorical(df["model_prefix"], categories=MODEL_ORDER, ordered=True)
    return df.sort_values(["domain", "model_prefix"]).reset_index(drop=True)


def load_raw_seed_results(results_dir: Path) -> pd.DataFrame:
    rows = []
    for domain in DOMAIN_ORDER:
        for model in MODEL_ORDER:
            for seed in SEEDS:
                path = results_dir / f"{model}_{domain}_seed{seed}.yaml"
                if not path.exists():
                    logger.warning("File tidak ditemukan, dilewati: %s", path)
                    continue
                with open(path) as f:
                    d = yaml.safe_load(f)
                rows.append({
                    "domain": domain,
                    "model_prefix": model,
                    "model_name": d.get("model_name", model),
                    "seed": seed,
                    "rmse": d["rmse"],
                    "mae": d["mae"],
                    "n_test_samples": d.get("n_test_samples"),
                    "precision_5": d.get("precision_at_k", {}).get(5),
                    "precision_10": d.get("precision_at_k", {}).get(10),
                    "precision_20": d.get("precision_at_k", {}).get(20),
                    "recall_5": d.get("recall_at_k", {}).get(5),
                    "recall_10": d.get("recall_at_k", {}).get(10),
                    "recall_20": d.get("recall_at_k", {}).get(20),
                    "ndcg_5": d.get("ndcg_at_k", {}).get(5),
                    "ndcg_10": d.get("ndcg_at_k", {}).get(10),
                    "ndcg_20": d.get("ndcg_at_k", {}).get(20),
                    "source_file": path.name,
                })
    df = pd.DataFrame(rows)
    df["domain"] = pd.Categorical(df["domain"], categories=DOMAIN_ORDER, ordered=True)
    df["model_prefix"] = pd.Categorical(df["model_prefix"], categories=MODEL_ORDER, ordered=True)
    return df.sort_values(["domain", "model_prefix", "seed"]).reset_index(drop=True)


def compute_significance(results_dir: Path) -> pd.DataFrame:
    """Hitung ulang uji Wilcoxon in-process dari predictions_*.csv (per-sample
    squared_error) via fungsi yang sudah ada di src/evaluation/metrics.py --
    TIDAK bergantung pada file significance_*.csv historis (ambigu utk
    domain restoran, lihat plan)."""
    rows = []
    for domain in DOMAIN_ORDER:
        for model_b in COMPARISON_MODELS:
            for seed in SEEDS:
                path_a = results_dir / f"predictions_baseline_reimpl_{domain}_seed{seed}.csv"
                path_b = results_dir / f"predictions_{model_b}_{domain}_seed{seed}.csv"
                if not path_a.exists() or not path_b.exists():
                    logger.warning("Predictions tidak lengkap utk %s/%s/seed%s, dilewati", domain, model_b, seed)
                    continue
                df_a = pd.read_csv(path_a)[["review_id", "squared_error"]].rename(columns={"squared_error": "se_a"})
                df_b = pd.read_csv(path_b)[["review_id", "squared_error"]].rename(columns={"squared_error": "se_b"})
                merged = df_a.merge(df_b, on="review_id", how="inner")
                statistic, p_value = significance_test(merged["se_a"].values, merged["se_b"].values, test="wilcoxon")
                rows.append({
                    "domain": domain,
                    "model_b": model_b,
                    "seed": seed,
                    "n_paired": len(merged),
                    "rmse_a": float(np.sqrt(merged["se_a"].mean())),
                    "rmse_b": float(np.sqrt(merged["se_b"].mean())),
                    "statistic": statistic,
                    "p_value": p_value,
                })
    df = pd.DataFrame(rows)
    df["domain"] = pd.Categorical(df["domain"], categories=DOMAIN_ORDER, ordered=True)
    df["model_b"] = pd.Categorical(df["model_b"], categories=COMPARISON_MODELS, ordered=True)
    return df.sort_values(["domain", "model_b", "seed"]).reset_index(drop=True)


def compute_stability_table(raw_df: pd.DataFrame) -> pd.DataFrame:
    """Replikasi metodologi Table IV manuskrip: SD penuh vs SD setelah
    mengeluarkan 1 seed dgn |z-score| terbesar, per model per domain."""
    rows = []
    for domain in DOMAIN_ORDER:
        for model in MODEL_ORDER:
            sub = raw_df[(raw_df["domain"] == domain) & (raw_df["model_prefix"] == model)].sort_values("seed")
            if len(sub) != 5:
                continue
            vals = sub["rmse"].to_numpy()
            seeds = sub["seed"].to_numpy()
            mean, std = vals.mean(), vals.std(ddof=1)
            z = (vals - mean) / std if std > 0 else np.zeros_like(vals)
            idx = int(np.argmax(np.abs(z)))
            mask = np.ones(len(vals), dtype=bool)
            mask[idx] = False
            std_excl = vals[mask].std(ddof=1) if mask.sum() > 1 else np.nan
            seed_map = dict(zip(seeds.tolist(), vals.tolist()))
            rows.append({
                "domain": domain,
                "model_prefix": model,
                "rmse_seed_42": seed_map.get(42),
                "rmse_seed_123": seed_map.get(123),
                "rmse_seed_456": seed_map.get(456),
                "rmse_seed_789": seed_map.get(789),
                "rmse_seed_1011": seed_map.get(1011),
                "sd_5_seed": std,
                "seed_paling_ekstrem": int(seeds[idx]),
                "z_score_ekstrem": float(z[idx]),
                "sd_tanpa_seed_ekstrem": float(std_excl) if std_excl == std_excl else None,
                "faktor_reduksi_sd": float(std / std_excl) if std_excl and std_excl > 0 else None,
            })
    df = pd.DataFrame(rows)
    df["domain"] = pd.Categorical(df["domain"], categories=DOMAIN_ORDER, ordered=True)
    df["model_prefix"] = pd.Categorical(df["model_prefix"], categories=MODEL_ORDER, ordered=True)
    return df.sort_values(["domain", "model_prefix"]).reset_index(drop=True)


def _get_nested(d: dict, dotted_path: str):
    cur = d
    for key in dotted_path.split("."):
        if not isinstance(cur, dict) or key not in cur:
            return None
        cur = cur[key]
    return cur


HYPERPARAM_ROWS = [
    ("DeepMF", "deepmf.embedding_dim"),
    ("DeepMF", "deepmf.hidden_layers"),
    ("DeepMF", "deepmf.dropout"),
    ("DeepMF", "deepmf.batch_size"),
    ("DeepMF", "deepmf.learning_rate"),
    ("DeepMF", "deepmf.negative_sampling_ratio"),
    ("CBF Clustering", "cbf_clustering.method"),
    ("CBF Clustering", "cbf_clustering.pca_components"),
    ("Fusion", "fusion_baseline.method"),
    ("Fusion", "fusion_baseline.nmf_components"),
    ("Fusion", "fusion_baseline.dt_max_depth"),
    ("Sentiment BERT", "sentiment_baseline.model_name"),
    ("Sentiment BERT", "sentiment_baseline.optimizer"),
    ("Sentiment BERT", "sentiment_baseline.learning_rate"),
    ("Sentiment BERT", "sentiment_baseline.epochs"),
    ("Split", "split.train_ratio"),
    ("Split", "split.val_ratio"),
    ("Split", "split.test_ratio"),
    ("Split", "data.min_reviews_per_user"),
    ("Split", "data.min_reviews_per_item"),
    ("CF Klasik", "classical_cf.item_knn.knn_k"),
    ("CF Klasik", "classical_cf.svd.svd_n_factors"),
    ("CF Klasik", "classical_cf.svd.svd_n_epochs"),
]


def extract_hyperparameters(results_dir: Path) -> pd.DataFrame:
    values_by_domain = {}
    for domain in DOMAIN_ORDER:
        path = results_dir / f"baseline_reimpl_{domain}_seed42.yaml"
        with open(path) as f:
            d = yaml.safe_load(f)
        values_by_domain[domain] = d.get("config_snapshot", {})

    rows = []
    for category, param_path in HYPERPARAM_ROWS:
        row = {"kategori": category, "parameter": param_path.split(".")[-1]}
        for domain in DOMAIN_ORDER:
            val = _get_nested(values_by_domain[domain], param_path)
            row[domain] = str(val) if val is not None else "-"
        rows.append(row)
    return pd.DataFrame(rows)


# ===========================================================================
# Styling helpers
# ===========================================================================

def style_header_row(ws, row_idx: int, n_cols: int, fill_hex: str = FILL_HEADER) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def autosize_columns(ws, max_width: int = 45) -> None:
    widths: dict[str, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None or cell.__class__.__name__ == "MergedCell":
                continue
            col_letter = get_column_letter(cell.column)
            widths[col_letter] = max(widths.get(col_letter, 0), len(str(cell.value)))
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = min(width + 2, max_width)


def set_tab_color(ws, hex_color: str) -> None:
    ws.sheet_properties.tabColor = hex_color


def write_dataframe(ws, df: pd.DataFrame, start_row: int, headers: list[str] | None = None) -> int:
    """Tulis df sbg tabel mulai dari start_row (1-indexed). Return baris
    terakhir yg ditulis (header + data)."""
    headers = headers or list(df.columns)
    for j, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=j, value=h)
    style_header_row(ws, start_row, len(headers))
    for i, (_, row) in enumerate(df.iterrows(), start=1):
        for j, col in enumerate(df.columns, start=1):
            val = row[col]
            if pd.isna(val):
                val = ""
            elif isinstance(val, (np.integer,)):
                val = int(val)
            elif isinstance(val, (np.floating,)):
                val = float(val)
            ws.cell(row=start_row + i, column=j, value=val)
    return start_row + len(df)


def apply_number_format(ws, col_letter: str, start_row: int, end_row: int, fmt: str) -> None:
    for r in range(start_row, end_row + 1):
        ws[f"{col_letter}{r}"].number_format = fmt


# ===========================================================================
# Sheet builders
# ===========================================================================

def build_sheet_0_sampul(wb: Workbook) -> None:
    ws = wb.create_sheet("0_Sampul")
    ws.column_dimensions["A"].width = 100
    r = 1
    ws.cell(row=r, column=1, value="Hasil Eksperimen A2-IRM: Aspect-Aware Integrated Representation Model")
    ws.cell(row=r, column=1).font = Font(bold=True, size=16)
    r += 1
    ws.cell(row=r, column=1, value=f"Digenerate: {date.today().isoformat()}")
    r += 1
    ws.cell(row=r, column=1, value="Cakupan: 3 domain x 7 model x 5 seed = 105 run eksperimen")
    r += 2

    ws.cell(row=r, column=1, value="CARA MEMBACA FILE INI").font = Font(bold=True, size=13)
    r += 1
    ws.cell(row=r, column=1,
            value="Untuk kesimpulan, cukup baca Sheet 1-3 (tab warna HIJAU). "
                  "Sheet 4 ke bawah (tab warna ABU-ABU) adalah lampiran teknis, "
                  "tersedia kalau ingin menelusuri detail atau ada pertanyaan verifikasi.")
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r].height = 40
    r += 2

    ws.cell(row=r, column=1, value="DAFTAR ISI").font = Font(bold=True, size=13)
    r += 1
    toc = [
        ("1_Ringkasan_Eksekutif", "[WAJIB] Temuan utama + chart headline"),
        ("2_Hasil_Utama_RMSE_MAE", "[WAJIB] Tabel RMSE/MAE lengkap 7 model x 3 domain"),
        ("3_Hybrid_vs_CF_Klasik", "[WAJIB] Bukti model hybrid mengungguli CF klasik"),
        ("4_Signifikansi_Statistik", "[Lampiran] Detail uji Wilcoxon per seed"),
        ("5_Stabilitas_Variansi", "[Lampiran] Analisis stabilitas antar-seed"),
        ("6_Data_Mentah_PerSeed", "[Lampiran] 105 baris data mentah per-seed"),
        ("7_Dataset_dan_Hyperparameter", "[Lampiran] Karakteristik dataset & hyperparameter"),
    ]
    for sheet_name, desc in toc:
        cell = ws.cell(row=r, column=1, value=f"-> {sheet_name}: {desc}")
        cell.hyperlink = f"#'{sheet_name}'!A1"
        cell.font = Font(color="0563C1", underline="single")
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Sumber data: checkpoints/results/ (lihat manuscript/A2-IRM_manuscript_draft.md "
                                    "utk metodologi lengkap dan draft manuskrip akademik).")
    ws.cell(row=r, column=1).font = Font(italic=True, size=9, color="595959")


def build_sheet_1_ringkasan(wb: Workbook, manuscript_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("1_Ringkasan_Eksekutif")
    set_tab_color(ws, FILL_TIER_WAJIB)
    r = 1
    ws.cell(row=r, column=1, value="Ringkasan Eksekutif").font = Font(bold=True, size=14)
    r += 2

    ws.cell(row=r, column=1, value="Temuan Utama").font = Font(bold=True, size=12)
    r += 1
    bullets = [
        "1. Model hybrid (DeepMF + Content-Based Filtering + Fusion) mengungguli CF klasik "
        "(Item-KNN/SVD) di ketiga domain -- RMSE turun 29-46%.",
        "2. Sentimen ber-aspek yang dirata-ratakan (ABSA-Mean / Confidence-Weighted Mean) justru "
        "MEMPERBURUK akurasi 12,9-21,7% dibanding sentimen global -- meratakan nuansa antar-aspek "
        "menghilangkan informasi, bukan menambahnya.",
        "3. Menyimpan skor tiap aspek + confidence sbg fitur terpisah (ABSA - Concat + Confidence) "
        "adalah SATU-SATUNYA varian yang konsisten mengalahkan baseline di ketiga domain "
        "(RMSE turun 2,0-3,2%), sekaligus varian paling stabil (variansi antar-seed 4-10x lebih rendah).",
        "4. Pola ini konsisten di 3 domain yang sangat berbeda karakternya (restoran/e-commerce/hotel, "
        "cakupan kata kunci aspek 45-96%) -- bukti kuat bahwa temuan ini bersifat umum, bukan kebetulan 1 dataset.",
    ]
    for b in bullets:
        cell = ws.cell(row=r, column=1, value=b)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 32
        r += 1
    r += 1

    # tabel ringkas per domain
    ws.cell(row=r, column=1, value="Ringkasan per Domain").font = Font(bold=True, size=12)
    r += 1
    summary_rows = []
    for domain in DOMAIN_ORDER:
        dom_df = manuscript_df[manuscript_df["domain"] == domain]
        core_df = dom_df[dom_df["model_prefix"].isin(CORE_MODELS)]
        best = core_df.loc[core_df["rmse_mean"].idxmin()]
        most_stable = core_df.loc[core_df["rmse_std"].idxmin()]
        item_knn_rmse = dom_df.loc[dom_df["model_prefix"] == "classical_cf_item_knn", "rmse_mean"].iloc[0]
        reduksi_pct = (item_knn_rmse - best["rmse_mean"]) / item_knn_rmse * 100
        summary_rows.append({
            "Domain": DOMAIN_LABEL[domain],
            "Model Terbaik (RMSE)": MODEL_LABEL[best["model_prefix"]],
            "RMSE": round(float(best["rmse_mean"]), 4),
            "Reduksi vs Item-KNN (%)": round(float(reduksi_pct), 1),
            "Model Paling Stabil (SD terendah)": MODEL_LABEL[most_stable["model_prefix"]],
        })
    summary_df = pd.DataFrame(summary_rows)
    table_start = r
    table_end = write_dataframe(ws, summary_df, r)
    apply_number_format(ws, "C", table_start + 1, table_end, "0.0000")
    apply_number_format(ws, "D", table_start + 1, table_end, "0.0")
    r = table_end + 2

    # data pendukung chart headline: baseline vs concat+confidence, 3 domain
    chart_start = r
    ws.cell(row=r, column=1, value="(Data pendukung grafik -- jangan diedit)").font = Font(italic=True, size=8, color="A6A6A6")
    r += 1
    chart_header_row = r
    ws.cell(row=r, column=1, value="Domain")
    ws.cell(row=r, column=2, value="Baseline")
    ws.cell(row=r, column=3, value="ABSA Concat+Confidence")
    r += 1
    chart_data_start = r
    for domain in DOMAIN_ORDER:
        dom_df = manuscript_df[manuscript_df["domain"] == domain]
        baseline_rmse = dom_df.loc[dom_df["model_prefix"] == "baseline_reimpl", "rmse_mean"].iloc[0]
        cc_rmse = dom_df.loc[dom_df["model_prefix"] == "absa_ablation_concat_confidence", "rmse_mean"].iloc[0]
        ws.cell(row=r, column=1, value=DOMAIN_LABEL[domain])
        ws.cell(row=r, column=2, value=round(float(baseline_rmse), 4))
        ws.cell(row=r, column=3, value=round(float(cc_rmse), 4))
        r += 1
    chart_data_end = r - 1

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "RMSE: Baseline vs Model Terbaik (ABSA Concat+Confidence)"
    chart.y_axis.title = "RMSE"
    chart.x_axis.title = "Domain"
    data = Reference(ws, min_col=2, max_col=3, min_row=chart_header_row, max_row=chart_data_end)
    cats = Reference(ws, min_col=1, min_row=chart_data_start, max_row=chart_data_end)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width, chart.height = 20, 11
    ws.add_chart(chart, f"F{table_start}")

    autosize_columns(ws, max_width=55)


def build_sheet_2_hasil_utama(wb: Workbook, manuscript_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("2_Hasil_Utama_RMSE_MAE")
    set_tab_color(ws, FILL_TIER_WAJIB)

    display = manuscript_df.copy()
    display["Domain"] = display["domain"].map(DOMAIN_LABEL)
    display["Model"] = display["model_prefix"].map(MODEL_LABEL)
    cols = [
        "Domain", "Model", "rmse_mean", "rmse_std", "mae_mean", "mae_std",
        "precision@5_mean", "precision@10_mean", "precision@20_mean",
        "recall@5_mean", "recall@10_mean", "recall@20_mean",
        "ndcg@5_mean", "ndcg@10_mean", "ndcg@20_mean",
        "p_value_combined_fisher", "n_significant_seeds",
    ]
    rename = {
        "rmse_mean": "RMSE_Rata2", "rmse_std": "RMSE_SD",
        "mae_mean": "MAE_Rata2", "mae_std": "MAE_SD",
        "precision@5_mean": "Precision@5", "precision@10_mean": "Precision@10", "precision@20_mean": "Precision@20",
        "recall@5_mean": "Recall@5", "recall@10_mean": "Recall@10", "recall@20_mean": "Recall@20",
        "ndcg@5_mean": "NDCG@5", "ndcg@10_mean": "NDCG@10", "ndcg@20_mean": "NDCG@20",
        "p_value_combined_fisher": "P_Value_Fisher_Gabungan", "n_significant_seeds": "Jumlah_Seed_Signifikan",
    }
    table_df = display[cols].rename(columns=rename)

    r = 1
    ws.cell(row=r, column=1,
            value="Catatan: metrik Precision/Recall/NDCG memakai protokol candidate-set terbatas "
                  "(bukan full-catalog) -- nilainya cenderung mendekati batas atas dan kurang "
                  "diskriminatif antar-model. RMSE/MAE adalah metrik utama yang jadi dasar kesimpulan.")
    ws.cell(row=r, column=1).font = Font(italic=True, size=9, color="595959")
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r].height = 28
    r += 2

    table_start = r
    table_end = write_dataframe(ws, table_df, r)
    apply_number_format(ws, "C", table_start + 1, table_end, "0.0000")
    apply_number_format(ws, "D", table_start + 1, table_end, "0.0000")
    apply_number_format(ws, "E", table_start + 1, table_end, "0.0000")
    apply_number_format(ws, "F", table_start + 1, table_end, "0.0000")
    for col_letter in ["G", "H", "I", "J", "K", "L", "M", "N", "O"]:
        apply_number_format(ws, col_letter, table_start + 1, table_end, "0.0000")
    apply_number_format(ws, "P", table_start + 1, table_end, "0.00E+00")

    ws.freeze_panes = f"A{table_start + 1}"
    ws.auto_filter.ref = f"A{table_start}:Q{table_end}"

    # pivot data utk chart: RMSE per model x domain
    r = table_end + 2
    ws.cell(row=r, column=1, value="(Data pendukung grafik -- jangan diedit)").font = Font(italic=True, size=8, color="A6A6A6")
    r += 1
    pivot = manuscript_df.pivot(index="model_prefix", columns="domain", values="rmse_mean")
    pivot = pivot.reindex(MODEL_ORDER)
    chart_header_row = r
    ws.cell(row=r, column=1, value="Model")
    for j, domain in enumerate(DOMAIN_ORDER, start=2):
        ws.cell(row=r, column=j, value=DOMAIN_LABEL[domain])
    r += 1
    chart_data_start = r
    for model in MODEL_ORDER:
        ws.cell(row=r, column=1, value=MODEL_LABEL[model])
        for j, domain in enumerate(DOMAIN_ORDER, start=2):
            ws.cell(row=r, column=j, value=round(float(pivot.loc[model, domain]), 4))
        r += 1
    chart_data_end = r - 1

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "RMSE per Model per Domain"
    chart.y_axis.title = "RMSE"
    chart.x_axis.title = "Model"
    data = Reference(ws, min_col=2, max_col=4, min_row=chart_header_row, max_row=chart_data_end)
    cats = Reference(ws, min_col=1, min_row=chart_data_start, max_row=chart_data_end)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width, chart.height = 26, 12
    ws.add_chart(chart, f"A{chart_data_end + 2}")

    autosize_columns(ws)


def build_sheet_3_hybrid_vs_cf(wb: Workbook, manuscript_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("3_Hybrid_vs_CF_Klasik")
    set_tab_color(ws, FILL_TIER_WAJIB)

    rows = []
    for domain in DOMAIN_ORDER:
        dom_df = manuscript_df[manuscript_df["domain"] == domain].set_index("model_prefix")
        rmse_hybrid = dom_df.loc["baseline_reimpl", "rmse_mean"]
        rmse_knn = dom_df.loc["classical_cf_item_knn", "rmse_mean"]
        rmse_svd = dom_df.loc["classical_cf_svd", "rmse_mean"]
        n_sig_knn = dom_df.loc["classical_cf_item_knn", "n_significant_seeds"]
        rows.append({
            "Domain": DOMAIN_LABEL[domain],
            "RMSE_Hybrid": round(float(rmse_hybrid), 4),
            "RMSE_ItemKNN": round(float(rmse_knn), 4),
            "RMSE_SVD": round(float(rmse_svd), 4),
            "Reduksi_vs_ItemKNN_%": round(float((rmse_knn - rmse_hybrid) / rmse_knn * 100), 1),
            "Reduksi_vs_SVD_%": round(float((rmse_svd - rmse_hybrid) / rmse_svd * 100), 1),
            "Signifikan": n_sig_knn,
        })
    table_df = pd.DataFrame(rows)

    r = 1
    ws.cell(row=r, column=1, value="Perbandingan Model Hybrid vs Collaborative Filtering Klasik").font = Font(bold=True, size=13)
    r += 2
    table_start = r
    table_end = write_dataframe(ws, table_df, r)
    for col_letter in ["B", "C", "D"]:
        apply_number_format(ws, col_letter, table_start + 1, table_end, "0.0000")
    for col_letter in ["E", "F"]:
        apply_number_format(ws, col_letter, table_start + 1, table_end, "0.0")

    r = table_end + 2
    ws.cell(row=r, column=1, value="(Data pendukung grafik -- jangan diedit)").font = Font(italic=True, size=8, color="A6A6A6")
    r += 1
    chart_header_row = r
    ws.cell(row=r, column=1, value="Domain")
    ws.cell(row=r, column=2, value="Hybrid")
    ws.cell(row=r, column=3, value="Item-KNN")
    ws.cell(row=r, column=4, value="SVD")
    r += 1
    chart_data_start = r
    for domain in DOMAIN_ORDER:
        dom_df = manuscript_df[manuscript_df["domain"] == domain].set_index("model_prefix")
        ws.cell(row=r, column=1, value=DOMAIN_LABEL[domain])
        ws.cell(row=r, column=2, value=round(float(dom_df.loc["baseline_reimpl", "rmse_mean"]), 4))
        ws.cell(row=r, column=3, value=round(float(dom_df.loc["classical_cf_item_knn", "rmse_mean"]), 4))
        ws.cell(row=r, column=4, value=round(float(dom_df.loc["classical_cf_svd", "rmse_mean"]), 4))
        r += 1
    chart_data_end = r - 1

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "RMSE: Hybrid vs Item-KNN vs SVD"
    chart.y_axis.title = "RMSE"
    chart.x_axis.title = "Domain"
    data = Reference(ws, min_col=2, max_col=4, min_row=chart_header_row, max_row=chart_data_end)
    cats = Reference(ws, min_col=1, min_row=chart_data_start, max_row=chart_data_end)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width, chart.height = 20, 11
    ws.add_chart(chart, f"H{table_start}")

    autosize_columns(ws)


def build_sheet_4_signifikansi(wb: Workbook, sig_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("4_Signifikansi_Statistik")
    set_tab_color(ws, FILL_TIER_LAMPIRAN)

    display = sig_df.copy()
    display["Domain"] = display["domain"].map(DOMAIN_LABEL)
    display["Model_Pembanding"] = display["model_b"].map(MODEL_LABEL)
    display["P_Value_Ringkas"] = display["p_value"].apply(lambda p: "<0.001" if p < 0.001 else f"{p:.4f}")
    display["Signifikan"] = display["p_value"].apply(lambda p: "Ya" if p < 0.05 else "Tidak")
    cols = ["Domain", "Model_Pembanding", "seed", "n_paired", "rmse_a", "rmse_b", "statistic", "p_value",
            "P_Value_Ringkas", "Signifikan"]
    rename = {
        "seed": "Seed", "n_paired": "N_Berpasangan", "rmse_a": "RMSE_Baseline", "rmse_b": "RMSE_Pembanding",
        "statistic": "Statistik_Wilcoxon", "p_value": "P_Value",
    }
    table_df = display[cols].rename(columns=rename)

    r = 1
    ws.cell(row=r, column=1, value="Signifikansi Statistik (Uji Wilcoxon Berpasangan per Seed)").font = Font(bold=True, size=13)
    r += 1
    ws.cell(row=r, column=1,
            value="Setiap baris = 1 perbandingan model vs baseline pada 1 seed, dihitung dari squared-error "
                  "per-sampel berpasangan pada test set yang identik.")
    ws.cell(row=r, column=1).font = Font(italic=True, size=9, color="595959")
    r += 2

    table_start = r
    table_end = write_dataframe(ws, table_df, r)
    apply_number_format(ws, "E", table_start + 1, table_end, "0.0000")
    apply_number_format(ws, "F", table_start + 1, table_end, "0.0000")
    apply_number_format(ws, "G", table_start + 1, table_end, "0.0000")
    apply_number_format(ws, "H", table_start + 1, table_end, "0.00E+00")

    green_fill = PatternFill(start_color=FILL_GREEN_SIG, end_color=FILL_GREEN_SIG, fill_type="solid")
    ws.conditional_formatting.add(
        f"H{table_start + 1}:H{table_end}",
        CellIsRule(operator="lessThan", formula=["0.05"], fill=green_fill),
    )
    ws.freeze_panes = f"A{table_start + 1}"
    ws.auto_filter.ref = f"A{table_start}:J{table_end}"
    autosize_columns(ws)


def build_sheet_5_stabilitas(wb: Workbook, stability_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("5_Stabilitas_Variansi")
    set_tab_color(ws, FILL_TIER_LAMPIRAN)

    display = stability_df.copy()
    display["Domain"] = display["domain"].map(DOMAIN_LABEL)
    display["Model"] = display["model_prefix"].map(MODEL_LABEL)
    cols = ["Domain", "Model", "rmse_seed_42", "rmse_seed_123", "rmse_seed_456", "rmse_seed_789", "rmse_seed_1011",
            "sd_5_seed", "seed_paling_ekstrem", "z_score_ekstrem", "sd_tanpa_seed_ekstrem", "faktor_reduksi_sd"]
    rename = {
        "rmse_seed_42": "RMSE_seed42", "rmse_seed_123": "RMSE_seed123", "rmse_seed_456": "RMSE_seed456",
        "rmse_seed_789": "RMSE_seed789", "rmse_seed_1011": "RMSE_seed1011",
        "sd_5_seed": "SD_5_Seed", "seed_paling_ekstrem": "Seed_Paling_Ekstrem",
        "z_score_ekstrem": "Z_Score_Ekstrem", "sd_tanpa_seed_ekstrem": "SD_Tanpa_Seed_Ekstrem",
        "faktor_reduksi_sd": "Faktor_Reduksi_SD",
    }
    table_df = display[cols].rename(columns=rename)

    r = 1
    ws.cell(row=r, column=1, value="Stabilitas & Variansi Antar-Seed").font = Font(bold=True, size=13)
    r += 1
    ws.cell(row=r, column=1,
            value="Catatan: 'Seed_Paling_Ekstrem' & 'Faktor_Reduksi_SD' menunjukkan seberapa besar 1 seed "
                  "tunggal mempengaruhi SD yang dilaporkan -- baseline domain Hotel (TripAdvisor) menunjukkan "
                  "Faktor_Reduksi_SD ~6.8x pada seed 123, jauh di atas domain lain (~1.3-1.5x), artinya SD "
                  "tinggi baseline di domain itu sebagian besar didorong 1 seed, bukan noise merata "
                  "(lihat manuscript/A2-IRM_manuscript_draft.md Section IV-C utk pembahasan lengkap).")
    ws.cell(row=r, column=1).font = Font(italic=True, size=9, color="595959")
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r].height = 40
    r += 2

    table_start = r
    table_end = write_dataframe(ws, table_df, r)
    for col_letter in ["C", "D", "E", "F", "G", "H", "J", "K"]:
        apply_number_format(ws, col_letter, table_start + 1, table_end, "0.0000")
    apply_number_format(ws, "L", table_start + 1, table_end, "0.0")

    ws.freeze_panes = f"A{table_start + 1}"
    ws.auto_filter.ref = f"A{table_start}:L{table_end}"

    # line chart per domain: baseline vs concat+confidence across seed
    r = table_end + 2
    ws.cell(row=r, column=1, value="(Data pendukung grafik -- jangan diedit)").font = Font(italic=True, size=8, color="A6A6A6")
    r += 1
    chart_anchor_row = r
    for domain in DOMAIN_ORDER:
        dom_df = stability_df[stability_df["domain"] == domain].set_index("model_prefix")
        baseline_row = dom_df.loc["baseline_reimpl"]
        cc_row = dom_df.loc["absa_ablation_concat_confidence"]
        header_row = r
        ws.cell(row=r, column=1, value="Seed")
        ws.cell(row=r, column=2, value="Baseline")
        ws.cell(row=r, column=3, value="ABSA Concat+Confidence")
        r += 1
        data_start = r
        for seed in SEEDS:
            ws.cell(row=r, column=1, value=seed)
            ws.cell(row=r, column=2, value=round(float(baseline_row[f"rmse_seed_{seed}"]), 4))
            ws.cell(row=r, column=3, value=round(float(cc_row[f"rmse_seed_{seed}"]), 4))
            r += 1
        data_end = r - 1

        chart = LineChart()
        chart.title = f"RMSE per Seed -- {DOMAIN_LABEL[domain]}"
        chart.y_axis.title = "RMSE"
        chart.x_axis.title = "Seed"
        data_ref = Reference(ws, min_col=2, max_col=3, min_row=header_row, max_row=data_end)
        cats_ref = Reference(ws, min_col=1, min_row=data_start, max_row=data_end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.width, chart.height = 16, 8
        ws.add_chart(chart, f"F{chart_anchor_row + DOMAIN_ORDER.index(domain) * 17}")
        r += 2

    autosize_columns(ws)


def build_sheet_6_data_mentah(wb: Workbook, raw_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("6_Data_Mentah_PerSeed")
    set_tab_color(ws, FILL_TIER_LAMPIRAN)

    display = raw_df.copy()
    display["Domain"] = display["domain"].map(DOMAIN_LABEL)
    display["Model"] = display["model_prefix"].map(MODEL_LABEL)
    cols = ["Domain", "Model", "seed", "rmse", "mae", "n_test_samples",
            "precision_5", "precision_10", "precision_20",
            "recall_5", "recall_10", "recall_20",
            "ndcg_5", "ndcg_10", "ndcg_20", "source_file"]
    rename = {
        "seed": "Seed", "rmse": "RMSE", "mae": "MAE", "n_test_samples": "N_Test_Samples",
        "precision_5": "Precision@5", "precision_10": "Precision@10", "precision_20": "Precision@20",
        "recall_5": "Recall@5", "recall_10": "Recall@10", "recall_20": "Recall@20",
        "ndcg_5": "NDCG@5", "ndcg_10": "NDCG@10", "ndcg_20": "NDCG@20", "source_file": "File_Sumber",
    }
    table_df = display[cols].rename(columns=rename)

    r = 1
    ws.cell(row=r, column=1, value="Data Mentah Per-Seed (105 Run Eksperimen)").font = Font(bold=True, size=13)
    r += 2
    table_start = r
    table_end = write_dataframe(ws, table_df, r)
    for col_letter in ["D", "E", "G", "H", "I", "J", "K", "L", "M", "N", "O"]:
        apply_number_format(ws, col_letter, table_start + 1, table_end, "0.0000")
    apply_number_format(ws, "F", table_start + 1, table_end, "#,##0")

    ws.freeze_panes = f"C{table_start + 1}"
    ws.auto_filter.ref = f"A{table_start}:P{table_end}"
    autosize_columns(ws)


def build_sheet_7_dataset_hyperparam(wb: Workbook, hp_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("7_Dataset_dan_Hyperparameter")
    set_tab_color(ws, FILL_TIER_LAMPIRAN)

    r = 1
    ws.cell(row=r, column=1, value="A. Karakteristik Dataset").font = Font(bold=True, size=13)
    r += 2
    dataset_rows = []
    for domain in DOMAIN_ORDER:
        s = DATASET_STATS[domain]
        dataset_rows.append({
            "Domain": DOMAIN_LABEL[domain],
            "Jumlah_Ulasan": s["reviews"],
            "Jumlah_User": s["users"],
            "Jumlah_Item": s["items"],
            "Sparsity": s["sparsity"],
            "Rating_Rata2": s["rating"],
            "Ukuran_Test_Set": s["test_size"],
        })
    dataset_df = pd.DataFrame(dataset_rows)
    table_start = r
    table_end = write_dataframe(ws, dataset_df, r)
    for col_letter in ["B", "C", "D", "G"]:
        apply_number_format(ws, col_letter, table_start + 1, table_end, "#,##0")
    apply_number_format(ws, "E", table_start + 1, table_end, "0.00%")
    apply_number_format(ws, "F", table_start + 1, table_end, "0.00")
    r = table_end + 3

    ws.cell(row=r, column=1, value="B. Hyperparameter Model (dibandingkan lintas domain)").font = Font(bold=True, size=13)
    r += 1
    ws.cell(row=r, column=1,
            value="Hyperparameter arsitektur sengaja dibuat identik lintas domain (kecuali path data) "
                  "utk menjamin perbandingan yang adil.")
    ws.cell(row=r, column=1).font = Font(italic=True, size=9, color="595959")
    r += 2
    hp_display = hp_df.rename(columns={
        "kategori": "Kategori", "parameter": "Parameter",
        "restaurant": "Nilai_Restoran", "amazon_electronics": "Nilai_Ecommerce", "tripadvisor_hotel": "Nilai_Hotel",
    })
    write_dataframe(ws, hp_display, r)

    autosize_columns(ws)


# ===========================================================================
# Main
# ===========================================================================

def main() -> None:
    parser = argparse.ArgumentParser(description="Generate laporan Excel bertahap hasil eksperimen A2-IRM")
    parser.add_argument("--results-dir", type=str, default="checkpoints/results")
    parser.add_argument("--output", type=str, default="manuscript/A2-IRM_hasil_eksperimen.xlsx")
    args = parser.parse_args()

    results_dir = Path(args.results_dir)

    logger.info("Memuat manuscript_table_*.csv ...")
    manuscript_df = load_manuscript_tables(results_dir)

    logger.info("Memuat & flatten 105 file YAML per-seed ...")
    raw_df = load_raw_seed_results(results_dir)

    logger.info("Menghitung ulang signifikansi Wilcoxon in-process (18 perbandingan x 5 seed) ...")
    sig_df = compute_significance(results_dir)

    logger.info("Menghitung tabel stabilitas/variansi ...")
    stability_df = compute_stability_table(raw_df)

    logger.info("Mengekstrak hyperparameter dari config_snapshot ...")
    hp_df = extract_hyperparameters(results_dir)

    logger.info("Menyusun workbook 8 sheet ...")
    wb = Workbook()
    del wb["Sheet"]
    build_sheet_0_sampul(wb)
    build_sheet_1_ringkasan(wb, manuscript_df)
    build_sheet_2_hasil_utama(wb, manuscript_df)
    build_sheet_3_hybrid_vs_cf(wb, manuscript_df)
    build_sheet_4_signifikansi(wb, sig_df)
    build_sheet_5_stabilitas(wb, stability_df)
    build_sheet_6_data_mentah(wb, raw_df)
    build_sheet_7_dataset_hyperparam(wb, hp_df)

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("Selesai. Laporan disimpan ke %s", output_path)


if __name__ == "__main__":
    main()
